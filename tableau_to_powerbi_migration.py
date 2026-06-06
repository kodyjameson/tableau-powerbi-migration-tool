import os, re, csv, zipfile, shutil
import xml.etree.ElementTree as ET

INPUT_FILE = os.path.expanduser(
    input("Drag the Tableau .twb or .twbx file here, then press Enter:\n")
    .strip()
    .strip("'")
    .strip('"')
    .replace("\\ ", " ")
)
OUTPUT_XLSX = "Tableau_Migration_Workbook.xlsx"
OUTPUT_FOLDER = "migration_csv_output"

def clean(v):
    if not v:
        return ""
    v = v.replace("[", "").replace("]", "").replace('"', "").strip()
    v = v.split(".")[-1]
    parts = v.split(":")
    bad = {"sum","none","usr","qk","nk","ok","pcto","Parameters","Sample - Superstore",
           "Measure Names","Multiple Values"}
    useful = [p for p in parts if p and p not in bad]
    v = useful[-1] if useful else v
    return "" if v in bad else v.strip()

def extract_twb(path):
    if path.lower().endswith(".twb"):
        return path
    if os.path.exists("extracted_twbx"):
        shutil.rmtree("extracted_twbx")
    os.makedirs("extracted_twbx")
    with zipfile.ZipFile(path, "r") as z:
        z.extractall("extracted_twbx")
    for root, dirs, files in os.walk("extracted_twbx"):
        for f in files:
            if f.lower().endswith(".twb"):
                return os.path.join(root, f)
    raise Exception("No TWB found inside TWBX.")

def write_csv(name, headers, rows):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    path = os.path.join(OUTPUT_FOLDER, name + ".csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)

twb = extract_twb(INPUT_FILE)
tree = ET.parse(twb)
root = tree.getroot()

calc_lookup = {}
calc_rows = []
dashboard_map = []
roadmap = []

for ds in root.findall(".//datasource"):
    ds_name = ds.get("name", "")
    for col in ds.findall(".//column"):
        calc = col.find("calculation")
        if calc is not None:
            raw = clean(col.get("name", ""))
            name = clean(col.get("caption", "")) or raw
            formula = calc.get("formula", "")
            calc_lookup[raw] = (name, formula)
            calc_lookup[name] = (name, formula)
            calc_rows.append([ds_name, name, formula])

for dash in root.findall(".//dashboard"):
    dash_name = dash.get("name", "")
    seen = set()
    for zone in dash.findall(".//zone"):
        ws = zone.get("name", "")
        if ws and ws not in seen:
            seen.add(ws)
            dashboard_map.append([dash_name, ws])

for ws in root.findall(".//worksheet"):
    ws_name = ws.get("name", "")
    found = set()
    for elem in ws.iter():
        for val in elem.attrib.values():
            for match in re.findall(r"\[[^\]]+\]", val):
                name = clean(match)
                if not name or name in found:
                    continue
                if name.startswith("Calculation_") and name not in calc_lookup:
                    continue
                found.add(name)
                if name in calc_lookup:
                    roadmap.append([ws_name, "Calculation", calc_lookup[name][0], calc_lookup[name][1]])
                else:
                    roadmap.append([ws_name, "Field", name, ""])

dash_to_ws = {}
for d, w in dashboard_map:
    dash_to_ws.setdefault(d, set()).add(w)

ws_to_deps = {}
for w, typ, dep, formula in roadmap:
    ws_to_deps.setdefault(w, []).append((typ, dep, formula))

full_roadmap = []
for d, worksheets in dash_to_ws.items():
    for w in sorted(worksheets):
        for typ, dep, formula in ws_to_deps.get(w, []):
            full_roadmap.append([d, w, typ, dep, formula])

summary = []
for d, worksheets in dash_to_ws.items():
    fields = set()
    calcs = set()
    for w in worksheets:
        for typ, dep, formula in ws_to_deps.get(w, []):
            if typ == "Field":
                fields.add(dep)
            elif typ == "Calculation":
                calcs.add(dep)
    summary.append([d, len(worksheets), len(fields), len(calcs)])

calc_deps = []
for ds, name, formula in calc_rows:
    fields = sorted(set(clean(x) for x in re.findall(r"\[[^\]]+\]", formula) if clean(x)))
    calc_deps.append([name, ", ".join(fields), formula])

sheets = {
    "Dashboard Map": (["Dashboard", "Worksheet"], dashboard_map),
    "Migration Roadmap": (["Dashboard", "Worksheet", "Dependency Type", "Dependency Name", "Formula"], full_roadmap),
    "Calculation Dictionary": (["Data Source", "Calculation Name", "Formula"], calc_rows),
    "Migration Summary": (["Dashboard", "Worksheet Count", "Field Count", "Calculation Count"], summary),
    "Calculation Dependencies": (["Calculation", "Fields Used", "Formula"], calc_deps),
}

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[letter].width = min(width + 3, 70)

    wb.save(OUTPUT_XLSX)
    print("Created:", OUTPUT_XLSX)

except ImportError:
    for sheet_name, (headers, rows) in sheets.items():
        write_csv(sheet_name.replace(" ", "_"), headers, rows)
    print("openpyxl not available. Created CSV folder:", OUTPUT_FOLDER)
