from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

INPUT_XLSX = input("Drag workbook to enhance, then press Enter:\n").strip().strip("'").strip('"').replace("\\ ", " ")
OUTPUT_XLSX = INPUT_XLSX.replace(".xlsx", "_Enhanced.xlsx")

wb = load_workbook(INPUT_XLSX)

roadmap = wb["Migration Roadmap"]

ws_stats = defaultdict(lambda: {"fields": set(), "calcs": set()})

for row in roadmap.iter_rows(min_row=2, values_only=True):
    dashboard, worksheet, dep_type, dep_name, formula = row

    if not worksheet or not dep_name:
        continue

    if dep_type == "Field":
        ws_stats[worksheet]["fields"].add(dep_name)

    if dep_type == "Calculation":
        ws_stats[worksheet]["calcs"].add(dep_name)

def complexity(score):
    if score <= 10:
        return "Low"
    if score <= 25:
        return "Medium"
    return "High"

def make_sheet(name, headers):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws

complexity_ws = make_sheet(
    "Worksheet Complexity",
    ["Worksheet", "Field Count", "Calculation Count", "Score", "Complexity"]
)

for worksheet, stats in sorted(ws_stats.items()):
    field_count = len(stats["fields"])
    calc_count = len(stats["calcs"])
    score = field_count + (calc_count * 3)

    complexity_ws.append([
        worksheet,
        field_count,
        calc_count,
        score,
        complexity(score)
    ])

build_ws = make_sheet(
    "Power BI Build Order",
    ["Step", "Build Item", "Purpose"]
)

build_steps = [
    [1, "Confirm data sources", "Identify and connect the required source tables/files."],
    [2, "Load source fields", "Bring in all required fields used by the Tableau workbook."],
    [3, "Recreate parameters", "Create slicers or disconnected tables for Tableau parameters."],
    [4, "Create base measures", "Build simple measures first, such as SUM(Sales)."],
    [5, "Create calculated measures", "Recreate Tableau calculated fields in DAX."],
    [6, "Validate dependencies", "Check that calculated fields have their required base fields."],
    [7, "Build visuals", "Recreate each worksheet as a Power BI visual."],
    [8, "Build dashboard page", "Arrange visuals to match the Tableau dashboard layout."],
    [9, "Validate totals and filters", "Compare Tableau and Power BI numbers."],
    [10, "Document gaps", "Log unsupported Tableau logic or manual rebuild needs."]
]

for step in build_steps:
    build_ws.append(step)

translate_ws = make_sheet(
    "Power BI Translation Hints",
    ["Tableau Pattern", "Power BI / DAX Direction", "Notes"]
)

hints = [
    ["SUM([Field])", "SUM(Table[Field])", "Basic aggregation."],
    ["AVG([Field])", "AVERAGE(Table[Field])", "Basic average."],
    ["IF ... THEN ... ELSE ... END", "IF(condition, true_result, false_result)", "May need nested IF or SWITCH."],
    ["CASE ... WHEN ... THEN ... END", "SWITCH()", "Usually maps well to SWITCH."],
    ["YEAR([Date])", "YEAR(Table[Date])", "Date function."],
    ["MONTH([Date])", "MONTH(Table[Date])", "Date function."],
    ["WINDOW_SUM", "DAX measure with filter context", "Usually needs manual review."],
    ["WINDOW_MAX", "DAX measure with filter context", "Usually needs manual review."],
    ["INDEX()", "RANKX or custom sort logic", "Manual review likely."],
    ["TOTAL()", "CALCULATE with ALL/ALLSELECTED", "Depends on Tableau view context."]
]

for hint in hints:
    translate_ws.append(hint)

for ws in wb.worksheets:
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 70)

wb.save(OUTPUT_XLSX)
print("Created:", OUTPUT_XLSX)
